from flask import render_template,send_file
from conexion import engine
from sqlalchemy import text


import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows



def cargarCalificaciones08Filtro(Periodo):
    try:
        query = text("""
            SELECT Matricula,Promedio,Comentarios,Periodo FROM foest08
                     WHERE Periodo = :Periodo;
        """)
        with engine.connect() as conn:
            rows = conn.execute(query,{"Periodo":Periodo}).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = f'''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                    <tr>
                       
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thexcel"><div>
                        <form action="/excel08filtro" method="post">
                            <input type="hidden" name="PeriodoSeleccionado" value="{Periodo}">
                            <button type="submit" class="btn btn-primary" id="excel">Generar excel</button>
                        </form>
                    </div></th>
                    </tr>
                    <tr>
                        <th>Matrícula</th>
                        <th>Promedio</th>
                        <th>Comentarios</th>
                        <th>Periodo</th>
                        <th><div>
                        <form action="/graficas08" method="post">
                            <button type="submit" class="btn btn-primary">Graficar todo</button>
                        </form>
                    </div></th>
                   
                    </tr>
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>

                    <td>
                        <form action="/calificacionCompleta08" method="post">
                            <input type="hidden" name="Matricula" value="{fila[0]}">
                            <button type="submit" class="btn btn-success btn-sm"id="calificaciones">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'

def cargarCalificaciones08():
    try:
        query = text("""
            SELECT Matricula,Promedio,Comentarios,Periodo FROM foest08;
        """)
        with engine.connect() as conn:
            rows = conn.execute(query).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = '''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                    <tr>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        
                        
                        
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thexcel"><div>
                        <form action="/excel08" method="post">
                            <button type="submit" class="btn btn-primary" id="excel">Generar excel</button>
                        </form>
                    </div></th>
                    </tr>
                    <tr>
                        <th>Matrícula</th>
                        <th>Promedio</th>
                        <th>Comentarios</th>
                        <th>Periodo</th>
                        <th><div>
                        <form action="/graficas08" method="post">
                            <button type="submit" class="btn btn-primary">Graficar todo</button>
                        </form>
                    </div></th>
                   
                    </tr>
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>

                    <td>
                        <form action="/calificacionCompleta08" method="post">
                            <input type="hidden" name="Matricula" value="{fila[0]}">
                            <button type="submit" class="btn btn-success btn-sm"id="calificaciones">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'


def cargarCalificaciones02():
    try:
        query = text("""
            SELECT foest02.Miembro, estudiante.Nombre1, estudiante.Nombre2, estudiante.ApellidoP, estudiante.ApellidoM,
                   foest02.PromedioActitud, foest02.PromedioDesarrollo, Periodo
            FROM estudiante 
            JOIN foest02 ON estudiante.Matricula = foest02.miembro
                     ;
        """)
        with engine.connect() as conn:
            rows = conn.execute(query).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = '''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                <tr>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thexcel"><div>
                        <form action="/excel02" method="post">
                            <button type="submit" class="btn btn-primary" id="excel">Excel periodo activo</button>
                        </form>
                    </div></th>
                    </tr>
                    <tr>
                        <th>Matrícula</th>
                        <th>Nombre</th>
                        <th>Segundo Nombre</th>
                        <th>Apellido Paterno</th>
                        <th>Apellido Materno</th>
                        <th>Prom. Actitud</th>
                        <th>Prom. Desarrollo</th>
                        <th>Periodo</th>
                        <th><div>
                        <form action="/graficas" method="post">
                            <button type="submit" class="btn btn-primary">Graficar periodo activo</button>
                        </form>
                    </div></th>
                    </tr>
                    
                    
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>
                    <td>{fila[4]}</td>
                    <td>{int(fila[5])}</td>
                    <td>{int(fila[6])}</td>
                    <td>{fila[7]}</td>
                    <td>
                        <form action="/calificacionCompleta" method="post">
                            <input type="hidden" name="Miembro" value="{fila[0]}">
                            <button type="submit" class="btn btn-success btn-sm" id="calificaciones">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'


def cargarCalificaciones02Filtro(Periodo):
    try:
        query = text("""
            SELECT foest02.Miembro, estudiante.Nombre1, estudiante.Nombre2, estudiante.ApellidoP, estudiante.ApellidoM,
                   foest02.PromedioActitud, foest02.PromedioDesarrollo, Periodo
            FROM estudiante 
            JOIN foest02 ON estudiante.Matricula = foest02.miembro
                     WHERE Periodo = :Periodo;
        """)
        with engine.connect() as conn:
            rows = conn.execute(query,{'Periodo':Periodo}).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = f'''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                    <tr>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thnborder"></th>
                        <th id="thexcel"><div>
                        <form action="/generarExcel02" method="post">
                            <input type="hidden" name="PeriodoSeleccionado" value="{Periodo}">
                            <button type="submit" class="btn btn-primary" id="excel">Generar excel</button>
                        </form>
                    </div></th>
                    </tr>
                    <tr>
                        <th>Matrícula</th>
                        <th>Nombre</th>
                        <th>Segundo Nombre</th>
                        <th>Apellido Paterno</th>
                        <th>Apellido Materno</th>
                        <th>Prom. Actitud</th>
                        <th>Prom. Desarrollo</th>
                        <th>Periodo</th>
                        <th><div>
                        <form action="/graficasfiltro" method="post">
                            <input type="hidden" name="Periodo" value="{Periodo}">
                            <button type="submit" class="btn btn-primary">Graficar</button>
                        </form>
                    </div></th>
                    </tr>
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>
                    <td>{fila[4]}</td>
                    <td>{int(fila[5])}</td>
                    <td>{int(fila[6])}</td>
                    <td>{fila[7]}</td>
                    <td>
                        <form action="/calificacionCompleta" method="post">
                            <input type="hidden" name="Miembro" value="{fila[0]}">
                            <button type="submit" class="btn btn-success btn-sm" id="calificaciones">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'


def cargarCalificaciones03():
    try:
        query = text("""
            SELECT DISTINCT proyecto.Nombre,
       calificacionproyectop1.Calificacion,
       calificacionproyectop2.Calificacion,
       calificacionproyectop3.Calificacion,
       equipos.Procedimiento
FROM proyecto
JOIN calificacionproyectop1 ON proyecto.Nombre = calificacionproyectop1.Proyecto
JOIN calificacionproyectop2 ON proyecto.Nombre = calificacionproyectop2.Proyecto
JOIN calificacionproyectop3 ON proyecto.Nombre = calificacionproyectop3.Proyecto
JOIN equipos ON proyecto.ProyectoID = equipos.Id_Proyecto;

        """)
        with engine.connect() as conn:
            rows = conn.execute(query).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = '''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                    <tr>
                        <th>Proyecto</th>
                        <th>Parcial 1</th>
                        <th>Parcial 2</th>
                        <th>Parcial 3</th>
                        <th>Procedimiento</th>
                        <th><div>
                        <form action="/graficar03" method="post">
                            <button type="submit" class="btn btn-primary">Graficar todo</button>
                        </form>
                    </div></th>
                    </tr>
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>
                    <td>{fila[4]}</td>
                 
                    <td>
                        <form action="/calificacionCompleta03" method="post">
                            <input type="hidden" name="Nombre" value="{fila[0]}">

                            <button type="submit" class="btn btn-success btn-sm">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'


def cargarCalificaciones03Filtro(Procedimiento):
    try:
        query = text("""
            SELECT DISTINCT proyecto.Nombre,
       calificacionproyectop1.Calificacion,
       calificacionproyectop2.Calificacion,
       calificacionproyectop3.Calificacion,
       equipos.Procedimiento
FROM proyecto
JOIN calificacionproyectop1 ON proyecto.Nombre = calificacionproyectop1.Proyecto
JOIN calificacionproyectop2 ON proyecto.Nombre = calificacionproyectop2.Proyecto
JOIN calificacionproyectop3 ON proyecto.Nombre = calificacionproyectop3.Proyecto
JOIN equipos ON proyecto.ProyectoID = equipos.Id_Proyecto
                     WHERE equipos.Procedimiento = :Procedimiento;

        """)
        with engine.connect() as conn:
            rows = conn.execute(query,{"Procedimiento":Procedimiento}).fetchall()  
            
            if not rows:
                return '<div class="alert alert-warning text-center">No hay datos disponibles</div>'

            resultado = f'''
            <div class="table-responsive">
            <table class="table table-bordered text-center align-middle">
                <thead class="table-light">
                    <tr>
                        <th>Proyecto</th>
                        <th>Parcial 1</th>
                        <th>Parcial 2</th>
                        <th>Parcial 3</th>
                        <th>Procedimiento</th>
                        <th><div>
                        <form action="/graficar03seleccionado" method="post">
                            <input type="hidden" name="Procedimiento" value="{Procedimiento}">
                            <button type="submit" class="btn btn-primary">Graficar</button>
                        </form>
                    </div></th>
                    </tr>
                </thead>
                <tbody>
            '''
            for fila in rows:
                resultado += f'''
                <tr>
                    <td>{fila[0]}</td>
                    <td>{fila[1]}</td>
                    <td>{fila[2]}</td>
                    <td>{fila[3]}</td>
                    <td>{fila[4]}</td>
                 
                    <td>
                        <form action="/calificacionCompleta03" method="post">
                            <input type="hidden" name="Nombre" value="{fila[0]}">

                            <button type="submit" class="btn btn-success btn-sm">Ver calif. completas</button>
                        </form>
                    </td>
                </tr>
                '''
            resultado += '''
                </tbody>
            </table>
            </div>
            '''
            return resultado
    except Exception as e:
        return f'<div class="alert alert-danger text-center">Error al cargar las calificaciones: {str(e)}</div>'




#Calificaciones completas de los formatos 2, 3 y 8

def Completa08(matricula):
    try:
        query = text("""
            SELECT * from foest08 where Matricula = :matricula
        """)

        with engine.connect() as conn:
            ok = conn.execute(query,{"matricula":matricula}).fetchall()

            
            if ok:
                for fila in ok:
                    Pregunta1 = fila[1]
                    Pregunta2 = fila[2]
                    Pregunta3 = fila[3]
                    Pregunta4 = fila[4]
                    Pregunta5 = fila[5]
                    Pregunta6 = fila[6]

                    Pregunta7 = fila[7]
                    Pregunta8 = fila[8]
                    Pregunta9 = fila[9]
                    Promedio= fila[10]
                    Veracidad = fila[11]

                    Comentarios = fila[12]
                    Matricula = fila[13]
                 
                   
    except Exception as e:
        return f'Error al cargar las calificaciones: {str(e)}'   
    return render_template('perfiles/Coordinacion/calificacionesCompletas08.html',Pregunta1 = Pregunta1,Pregunta2 = Pregunta2,Pregunta3 = Pregunta3,Pregunta4 = Pregunta4,Pregunta5 = Pregunta5,Pregunta6 = Pregunta6,Pregunta7 = Pregunta7,
                           Pregunta8 = Pregunta8,Pregunta9 = Pregunta9,Promedio = Promedio,Veracidad = Veracidad,Comentarios = Comentarios,Matricula = Matricula)

def Completa03(nombre):
    
    try:
        query = text("""
            SELECT calificacionproyectop1.*,calificacionproyectop2.*,calificacionproyectop3.*
FROM proyecto
JOIN calificacionproyectop1 ON proyecto.Nombre = calificacionproyectop1.Proyecto
JOIN calificacionproyectop2 ON proyecto.Nombre = calificacionproyectop2.Proyecto
JOIN calificacionproyectop3 ON proyecto.Nombre = calificacionproyectop3.Proyecto
WHERE proyecto.Nombre = :Nombre;
        """)

        with engine.connect() as conn:
            ok = conn.execute(query,{"Nombre":nombre}).fetchall()

            
            if ok:
                for fila in ok:
                    Antecedentes = fila[2]
                    Planteamiento = fila[3]
                    Justificacion = fila[4]
                    Objetivos = fila[5]
                    ObjetivosEspecificos = fila[6]
                    PromedioP1 = fila[7]

                    Marco = fila[10]
                    Metodologia = fila[11]
                    Cronograma = fila[12]
                    Desarrollo= fila[13]
                    PromedioP2 = fila[14]

                    Resultados = fila[17]
                    Conclusiones = fila[18]
                    Referencias = fila[19]
                    Anexos = fila[20]
                    PromedioP3 = fila[21]
                   
    except Exception as e:
        return f'Error al cargar las calificaciones: {str(e)}'   
    return render_template('perfiles/Coordinacion/calificacionesCompletas03.html',Antecedentes=Antecedentes,Planteamiento=Planteamiento,Justificacion=Justificacion,
                           Objetivos = Objetivos,ObjetivosEspecificos=ObjetivosEspecificos,PromedioP1 = PromedioP1,Marco = Marco,Metodologia = Metodologia,Cronograma = Cronograma,
                           Desarrollo = Desarrollo,PromedioP2 = PromedioP2,Resultados = Resultados,Conclusiones = Conclusiones,Referencias = Referencias,Anexos = Anexos,PromedioP3 = PromedioP3)

def Completa02(Miembro):
    try:
        query = text("""
            SELECT estudiante.Nombre1, estudiante.Nombre2, estudiante.ApellidoP, estudiante.ApellidoM, foest02.*
            FROM estudiante 
            JOIN foest02 ON estudiante.Matricula = foest02.Miembro
            WHERE foest02.Miembro = :miembro
        """)

        with engine.connect() as conn:
            ok = conn.execute(query, {'miembro': Miembro}).fetchall()

            
            if ok:
                for fila in ok:
                    
                    
                    Puntualidad = fila[11]
                    Responsabilidad = fila[12]
                    Etica = fila[13]
                    TomaDecisiones = fila[14]
                    Liderazgo = fila[15]
                    ExpresaIdeas= fila[16]
                    ComunicacionAsertiva = fila[17]
                    ResolucionSituaciones = fila[18]
                    ActitudFavorable = fila[19]
                    TrabajoEquipo = fila[20]
                    promedio_actitud = fila[21]
                    Estrategias = fila[22]
                    AccionesMejora = fila[23]
                    ProcesosOperacion = fila[24]
                    PlanteaSoluciones = fila[25]
                    RespondeNecesidades = fila[26]
                    CumpleTiempos = fila[27]
                    promedio_desarrollo = fila[28]
                    

                    promedio_actitud=int(promedio_actitud)
                    promedio_desarrollo=int(promedio_desarrollo)
    except Exception as e:
        return f'Error al cargar las calificaciones: {str(e)}'


                
    return render_template('perfiles/Coordinacion/calificacionCompleta.html',Puntualidad=Puntualidad, Responsabilidad=Responsabilidad, Etica=Etica, TomaDecisiones=TomaDecisiones, Liderazgo=Liderazgo,ExpresaIdeas=ExpresaIdeas,
                           ComunicacionAsertiva=ComunicacionAsertiva, ResolucionSituaciones=ResolucionSituaciones, ActitudFavorable=ActitudFavorable, TrabajoEquipo=TrabajoEquipo,Estrategias=Estrategias,
                           AccionesMejora=AccionesMejora, ProcesosOperacion=ProcesosOperacion, PlanteaSoluciones=PlanteaSoluciones, RespondeNecesidades=RespondeNecesidades, CumpleTiempos=CumpleTiempos,promedio_actitud=promedio_actitud,promedio_desarrollo=promedio_desarrollo)


# Cargar proyectos de coordinación para el parcial 3
def cargarproyectoscoordinacion():
    query = text("SELECT NombreProyecto FROM documentos WHERE Parcial = 'Parcial 3'")
    with engine.connect() as conn:
        ok = conn.execute(query)
        rows = ok.fetchall()

        if rows:
            tarjetas = ''
            total = len(rows)

            for row in rows:
                nombre = row[0]
                tarjetas += f'''
                <div class="col-md-4 col-sm-6 col-10 mb-4 d-flex justify-content-center">
                    <div class="card text-center shadow-sm proyecto-card">
                        <div class="card-body d-flex flex-column justify-content-between">
                            <h5 class="card-title fw-bold">{nombre}</h5>
                            <form action="/abrirproyectoruta" method="post">
                                <input type="hidden" name="nombre" value="{nombre}">
                                <button type="submit" class="btn btn-success mt-3">Abrir</button>
                            </form>
                        </div>
                    </div>
                </div>
                '''

            sobrantes = (3 - (total % 3)) % 3  

            for _ in range(sobrantes):
                tarjetas += '''
                <div class="col-md-4 col-sm-6 col-10 mb-4 d-flex justify-content-center">
                    <div class="proyecto-card invisible"></div>
                </div>
                '''

            return tarjetas
        else:
            return '<div class="alert alert-warning text-center">NADA POR MOSTRAR</div>'
        










def excel02filtrado(periodo):
    try:
        # Consulta SQL con parámetro :Periodo
        query = text("""
            SELECT Periodo,foest02.Miembro, estudiante.Nombre1, estudiante.Nombre2, 
                   estudiante.ApellidoP, estudiante.ApellidoM,estudiante.Grupo,foest02.Puntualidad,
                   foest02.Responsabilidad, foest02.Etica, foest02.TomaDecisiones,
                   foest02.Liderazgo, foest02.ExpresaIdeas, foest02.ComunicacionAsertiva,
                   foest02.ResolucionSituaciones, foest02.ActitudFavorable,foest02.TrabajoEnEquipo,
                   foest02.PromedioActitud,foest02.Estrategias, foest02.AccionesMejora,
                   foest02.ProcesosOperacion, foest02.PlanteaSoluciones,foest02.RespondeNecesidades,
                   foest02.CumpleTiempo,foest02.PromedioDesarrollo
            FROM estudiante 
            JOIN foest02 ON estudiante.Matricula = foest02.Miembro
            WHERE Periodo = :Periodo;
        """)

        # Ejecutar la consulta
        with engine.connect() as conn:
            result = conn.execute(query, {'Periodo': periodo})
            rows = result.fetchall()
            columns = result.keys()

        if not rows:
            return f"No se encontraron resultados para el periodo {periodo}", 404

        # Convertir a DataFrame y personalizar encabezados
        df = pd.DataFrame(rows, columns=columns)
        df.columns = [
            "Periodo","Matrícula", "Primer Nombre", "Segundo Nombre", "Apellido Paterno", "Apellido Materno","Grupo",
            "Puntualidad", "Responsabilidad", "Ética", "Toma de Decisiones", "Liderazgo",
            "Expresa Ideas", "Comunicación Asertiva", "Resolución de Situaciones",
            "Actitud Favorable", "Trabajo en Equipo",
            "Promedio de Actitud","Estrategias", "Acciones de Mejora",
            "Procesos de Operación", "Plantea Soluciones", "Responde Necesidades",
            "Cumple Tiempos",
            "Promedio de Desarrollo"
        ]

        # Crear archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = f"FO-EST-02 {periodo}"

        # Escribir encabezados y datos
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2


     # Colorear toda la columna P y W de amarillo (incluyendo encabezado)
        yellow_fill = PatternFill(start_color="FFFC00", end_color="FFFC00", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row[15].fill = yellow_fill  # Columna P (índice 15, empieza en 0)
            row[22].fill = yellow_fill  # Columna W (índice 22, empieza en 0)
            ws[1][15].font = Font(bold=True)  # Encabezado columna P
            ws[1][22].font = Font(bold=True)  # Encabezado columna W


        # Guardar el Excel en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar archivo al navegador
        return send_file(
            output,
            download_name=f"FO-EST-02 {periodo}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return f"Error al generar el Excel: {e}", 500


def excel08(periodo):
    try:
        # Consulta SQL con parámetro :Periodo
        query = text("""
SELECT foest08.Periodo,foest08.Matricula,estudiante.Nombre1,estudiante.Nombre2,estudiante.ApellidoP,estudiante.ApellidoM,estudiante.Grupo,foest08.Pregunta01,foest08.Pregunta02,foest08.Pregunta03,foest08.Pregunta04,foest08.Pregunta05,foest08.Pregunta06,foest08.Pregunta07,foest08.Pregunta08,foest08.Pregunta09,foest08.Promedio,foest08.Veracidad,foest08.Comentarios
                     from foest08
                     JOIN estudiante ON foest08.Matricula = estudiante.Matricula
                     WHERE foest08.periodo = :periodo;
        """)

        # Ejecutar la consulta
        with engine.connect() as conn:
            result = conn.execute(query,{"periodo":periodo})
            rows = result.fetchall()
            columns = result.keys()

        if not rows:
            return f"No se encontraron resultados para el periodo {periodo}", 404

        # Convertir a DataFrame y personalizar encabezados
        df = pd.DataFrame(rows, columns=columns)
        df.columns = [
            "Periodo","Matrícula", "Primer Nombre", "Segundo Nombre", "Apellido Paterno", "Apellido Materno","Grupo",
            "¿La coordinación de estancias y estadías ofrece un servicio de calidad?", "¿Consideras que el responsable de vinculación proporciona información útil para que los alumnos encuentren empresa dónde realizar su estancia/estadía?", "¿La plataforma para generar cartas de representación es eficiente y fácil de utilizar?", "¿La gestión de mi carta de representación fue oportuna?", "¿La asignación de tu asesor académico fue oportuna?",
            "¿La capacitación sobre la estructura y elaboración del proyecto de estancia/estadía recibida por parte del asesor académico fue adecuada?", "¿El asesor académico asistió constantemente a las asesorías programadas?", "¿El asesor académico te apoyó en la aclaración de dudas técnicas durante el desarrollo de tu proyecto?",
            "¿El asesor académico mantuvo contacto con el asesor empresarial para un adecuado desarrollo del proyecto de estancia/estadía?","Promedio", "Veracidad",
            "Comentarios"
        ]

        # Crear archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = f"FO-EST-08 {periodo}"

        # Escribir encabezados y datos
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length+2


     # Colorear toda la columna P y W de amarillo (incluyendo encabezado)
        yellow_fill = PatternFill(start_color="FFFC00", end_color="FFFC00", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row[15].fill = yellow_fill  # Columna P (índice 15, empieza en 0)
            ws[1][15].font = Font(bold=True)  # Encabezado columna P
            


        # Guardar el Excel en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar archivo al navegador
        return send_file(
            output,
            download_name=f"FO-EST-08 {periodo}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return f"Error al generar el Excel: {e}", 500




def excel07P(periodo,procedimiento):

    try:
        # Consulta SQL con parámetro :Periodo
        query = text("""SELECT 
	foest07.Periodo,
    estudiante.Grupo,
    GROUP_CONCAT(CONCAT(estudiante.Nombre1, ' ', estudiante.ApellidoP, ' ', estudiante.ApellidoM) SEPARATOR ', ') AS Integrantes,
    foest07.TituloProyecto,
    foest07.NombreEmpresa,
    estudiante.Carrera,
    foest07.Procedimiento,
    GROUP_CONCAT(DISTINCT
    TRIM(BOTH ', ' FROM CONCAT_WS(', ',
        NULLIF(TRIM(foest07.FuncionesPrioritarias), ''),
        NULLIF(TRIM(foest07.FuncionesPrioritarias2), ''),
        NULLIF(TRIM(foest07.FuncionesPrioritarias3), '')
    ))
    SEPARATOR ', '
) AS FuncionesPrioritarias,
    foest07.ResolvioNecesidad,
    foest07.InteresParticipacion,
    foest07.RealizaInvestigacion,
    foest07.DispuestoContratar,
    foest07.PorqueContratar,
    foest07.ApruebaEdicion,
    foest07.ClausulaEspecial
FROM estudiante
JOIN equipos ON estudiante.Matricula = equipos.Matricula
JOIN proyecto ON equipos.Id_Proyecto = proyecto.ProyectoID
JOIN foest07 ON proyecto.Nombre = foest07.TituloProyecto
WHERE foest07.Periodo = :periodo AND foest07.Procedimiento = :procedimiento
GROUP BY equipos.NoEquipo, foest07.TituloProyecto, estudiante.Grupo;""")

        # Ejecutar la consulta
        with engine.connect() as conn:
            result = conn.execute(query,{"periodo":periodo, "procedimiento":procedimiento})
            rows = result.fetchall()
            columns = result.keys()

        if not rows:
            return f"No se encontraron resultados para el periodo {periodo}", 404

        # Convertir a DataFrame y personalizar encabezados
        df = pd.DataFrame(rows, columns=columns)
        df.columns = [
            "Periodo","Grupo", "Integrantes", "Titulo del proyecto", "Nombre de la empresa", "Carrera","Procedimiento",
            "Funciones prioritarias", "El proyecto resolvió su necesidad", "Le interesaría participar en proyectos de investigación y desarrollo tecnológico,asesorados por la UPT",
            "¿Realiza investigación y desarrollo en su empresa?", "¿Estaria dispuesto a contratar a egresados de la UPT?", "¿Por qué?","Aprueba edición del proyecto por parte de la UPT",
            "Clausula especial de las condiciones de edición por parte del asesor empresarial en caso de existir"
        ]

        # Crear archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = f"FO-EST-07 {periodo} {procedimiento}"

        # Escribir encabezados y datos
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length+2


     # Colorear toda la columna P y W de amarillo (incluyendo encabezado)
       
        

        # Encuentra el índice de la columna "Funciones prioritarias"
        # Encuentra el índice de la columna "Funciones prioritarias"
       

                # Guardar el Excel en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar archivo al navegador
        return send_file(
            output,
            download_name=f"FO-EST-07 {periodo} {procedimiento}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return f"Error al generar el Excel: {e}", 500



def excel07(periodo):

    try:
        # Consulta SQL con parámetro :Periodo
        query = text("""SELECT 
	foest07.Periodo,
    estudiante.Grupo,
    GROUP_CONCAT(CONCAT(estudiante.Nombre1, ' ', estudiante.ApellidoP, ' ', estudiante.ApellidoM) SEPARATOR ', ') AS Integrantes,
    foest07.TituloProyecto,
    foest07.NombreEmpresa,
    estudiante.Carrera,
    foest07.Procedimiento,
    GROUP_CONCAT(DISTINCT
    TRIM(BOTH ', ' FROM CONCAT_WS(', ',
        NULLIF(TRIM(foest07.FuncionesPrioritarias), ''),
        NULLIF(TRIM(foest07.FuncionesPrioritarias2), ''),
        NULLIF(TRIM(foest07.FuncionesPrioritarias3), '')
    ))
    SEPARATOR ', '
) AS FuncionesPrioritarias,
    foest07.ResolvioNecesidad,
    foest07.InteresParticipacion,
    foest07.RealizaInvestigacion,
    foest07.DispuestoContratar,
    foest07.PorqueContratar,
    foest07.ApruebaEdicion,
    foest07.ClausulaEspecial
FROM estudiante
JOIN equipos ON estudiante.Matricula = equipos.Matricula
JOIN proyecto ON equipos.Id_Proyecto = proyecto.ProyectoID
JOIN foest07 ON proyecto.Nombre = foest07.TituloProyecto
WHERE foest07.Periodo = :periodo
GROUP BY equipos.NoEquipo, foest07.TituloProyecto, estudiante.Grupo;""")

        # Ejecutar la consulta
        with engine.connect() as conn:
            result = conn.execute(query,{"periodo":periodo})
            rows = result.fetchall()
            columns = result.keys()

        if not rows:
            return f"No se encontraron resultados para el periodo {periodo}", 404

        # Convertir a DataFrame y personalizar encabezados
        df = pd.DataFrame(rows, columns=columns)
        df.columns = [
            "Periodo","Grupo", "Integrantes", "Titulo del proyecto", "Nombre de la empresa", "Carrera","Procedimiento",
            "Funciones prioritarias", "El proyecto resolvió su necesidad", "Le interesaría participar en proyectos de investigación y desarrollo tecnológico,asesorados por la UPT",
            "¿Realiza investigación y desarrollo en su empresa?", "¿Estaria dispuesto a contratar a egresados de la UPT?", "¿Por qué?","Aprueba edición del proyecto por parte de la UPT",
            "Clausula especial de las condiciones de edición por parte del asesor empresarial en caso de existir"
        ]

        # Crear archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = f"FO-EST-07 {periodo}"

        # Escribir encabezados y datos
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if i == 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length+2


     # Colorear toda la columna P y W de amarillo (incluyendo encabezado)
       
        

        # Encuentra el índice de la columna "Funciones prioritarias"
        # Encuentra el índice de la columna "Funciones prioritarias"
       

                # Guardar el Excel en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Enviar archivo al navegador
        return send_file(
            output,
            download_name=f"FO-EST-07 {periodo}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return f"Error al generar el Excel: {e}", 500



